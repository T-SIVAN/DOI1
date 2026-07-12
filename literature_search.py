"""Federated scholarly literature search and export helpers.

The module deliberately keeps provider clients small and dependency-light so it
can be used by both a web application and command-line jobs.
"""

from __future__ import annotations

import abc
import concurrent.futures
import copy
import difflib
import io
import json
import re
import threading
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from typing import Any, Callable, Iterable, Mapping, Sequence


@dataclass
class PaperRecord:
    title: str
    source: str
    authors: list[str] = field(default_factory=list)
    year: int | None = None
    publication: str = ""
    volume: str = ""
    issue: str = ""
    pages: str = ""
    abstract: str = ""
    doi: str = ""
    pmid: str = ""
    pmcid: str = ""
    openalex_id: str = ""
    url: str = ""
    citations: int | None = None
    is_open_access: bool | None = None
    oa_url: str = ""
    oa_license: str = ""
    mesh_terms: list[str] = field(default_factory=list)
    keywords: list[str] = field(default_factory=list)
    publication_types: list[str] = field(default_factory=list)
    is_retracted: bool = False
    rank: int = 0
    score: float = 0.0
    sources: list[str] = field(default_factory=list)
    source_ids: dict[str, str] = field(default_factory=dict)
    source_ranks: dict[str, int] = field(default_factory=dict)
    field_sources: dict[str, str] = field(default_factory=dict)
    fetched_at: str = ""
    missing_fields: list[str] = field(default_factory=list)
    summary_ready: bool = False
    reference_ready: bool = False
    needs_enrich: bool = True
    quality_flags: list[str] = field(default_factory=list)


@dataclass
class SourceSearchResult:
    source: str
    records: list[PaperRecord] = field(default_factory=list)
    query: str = ""
    error: str | None = None
    elapsed_seconds: float = 0.0


@dataclass
class FederatedSearchResult:
    records: list[PaperRecord] = field(default_factory=list)
    source_results: dict[str, SourceSearchResult] = field(default_factory=dict)
    queries: list[str] = field(default_factory=list)
    original_query: str = ""
    source_queries: dict[str, str] = field(default_factory=dict)
    terms_en: list[str] = field(default_factory=list)
    terms_zh: list[str] = field(default_factory=list)
    raw_count: int = 0
    unique_count: int = 0
    duplicate_count: int = 0
    source_counts: dict[str, int] = field(default_factory=dict)
    elapsed_seconds: float = 0.0
    errors: dict[str, str] = field(default_factory=dict)


class SourceAdapter(abc.ABC):
    """Interface implemented by each literature provider."""

    name: str

    @abc.abstractmethod
    def search(
        self,
        query: str,
        start_year: int | None,
        end_year: int | None,
        limit: int,
        credentials: Mapping[str, str] | None = None,
    ) -> SourceSearchResult:
        raise NotImplementedError


class HTTPClient:
    """Small urllib client with bounded retries and Retry-After support."""

    def __init__(self, timeout: float = 25.0, retries: int = 1) -> None:
        self.timeout = timeout
        self.retries = retries

    def get_text(
        self,
        url: str,
        params: Mapping[str, Any] | None = None,
        headers: Mapping[str, str] | None = None,
    ) -> str:
        if params:
            encoded = urllib.parse.urlencode(
                {key: value for key, value in params.items() if value is not None},
                doseq=True,
            )
            url = f"{url}{'&' if '?' in url else '?'}{encoded}"
        request = urllib.request.Request(
            url,
            headers={"User-Agent": "FederatedLiteratureSearch/1.0", **(headers or {})},
        )
        for attempt in range(self.retries + 1):
            try:
                with urllib.request.urlopen(request, timeout=self.timeout) as response:
                    return response.read().decode(response.headers.get_content_charset() or "utf-8")
            except urllib.error.HTTPError as exc:
                if attempt >= self.retries or exc.code not in {429, 500, 502, 503, 504}:
                    raise
                delay = _retry_delay(exc.headers.get("Retry-After"), attempt)
            except (urllib.error.URLError, TimeoutError):
                if attempt >= self.retries:
                    raise
                delay = min(2**attempt, 8)
            time.sleep(delay)
        raise RuntimeError("HTTP retry loop exhausted")

    def get_json(
        self,
        url: str,
        params: Mapping[str, Any] | None = None,
        headers: Mapping[str, str] | None = None,
    ) -> Any:
        return json.loads(self.get_text(url, params=params, headers=headers))


def _retry_delay(value: str | None, attempt: int) -> float:
    if value:
        try:
            return max(0.0, min(float(value), 60.0))
        except ValueError:
            try:
                retry_at = parsedate_to_datetime(value)
                if retry_at.tzinfo is None:
                    retry_at = retry_at.replace(tzinfo=timezone.utc)
                seconds = (retry_at - datetime.now(timezone.utc)).total_seconds()
                return max(0.0, min(seconds, 60.0))
            except (TypeError, ValueError, OverflowError):
                pass
    return min(2**attempt, 8)


def _text(element: ET.Element | None) -> str:
    return "" if element is None else "".join(element.itertext()).strip()


def _clean_doi(value: Any) -> str:
    doi = str(value or "").strip()
    doi = re.sub(r"^(?:https?://(?:dx\.)?doi\.org/|doi:\s*)", "", doi, flags=re.I)
    return doi.rstrip(" .").lower()


def _clean_openalex_id(value: Any) -> str:
    identifier = str(value or "").strip().rstrip("/")
    if "/" in identifier:
        identifier = identifier.rsplit("/", 1)[-1]
    return identifier.upper()


def _year(value: Any) -> int | None:
    match = re.search(r"(?:18|19|20|21)\d{2}", str(value or ""))
    return int(match.group()) if match else None


def _strip_markup(value: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    return re.sub(r"\s+", " ", text).strip()


def _bounded_limit(limit: Any) -> int:
    try:
        value = int(limit)
    except (TypeError, ValueError):
        value = 20
    return max(1, min(value, 50))


_NCBI_RATE_LOCK = threading.Lock()
_NCBI_LAST_REQUEST = 0.0


def _wait_for_ncbi(api_key: str | None) -> None:
    """Apply NCBI's process-wide anonymous/API-key request interval."""
    global _NCBI_LAST_REQUEST
    interval = 0.1 if api_key else (1.0 / 3.0)
    with _NCBI_RATE_LOCK:
        now = time.monotonic()
        delay = interval - (now - _NCBI_LAST_REQUEST)
        if delay > 0:
            time.sleep(delay)
        _NCBI_LAST_REQUEST = time.monotonic()


class PubMedAdapter(SourceAdapter):
    name = "pubmed"
    _base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils"

    def __init__(self, http: HTTPClient | None = None) -> None:
        self.http = http or HTTPClient()

    def search(
        self,
        query: str,
        start_year: int | None,
        end_year: int | None,
        limit: int,
        credentials: Mapping[str, str] | None = None,
    ) -> SourceSearchResult:
        started = time.monotonic()
        credentials = credentials or {}
        limit = _bounded_limit(limit)
        bounded = _bounded_query(query, start_year, end_year, "pubmed")
        api_key = credentials.get("pubmed_api_key") or credentials.get("ncbi_api_key")
        common = {
            "api_key": api_key,
            "email": credentials.get("email"),
            "tool": credentials.get("ncbi_tool", "federated-literature-search"),
        }
        _wait_for_ncbi(api_key)
        payload = self.http.get_json(
            f"{self._base}/esearch.fcgi",
            {
                "db": "pubmed",
                "term": bounded,
                "retmode": "json",
                "retmax": limit,
                "sort": "relevance",
                **common,
            },
        )
        ids = [str(item) for item in payload.get("esearchresult", {}).get("idlist", [])]
        records: list[PaperRecord] = []
        if ids:
            _wait_for_ncbi(api_key)
            xml_text = self.http.get_text(
                f"{self._base}/efetch.fcgi",
                {
                    "db": "pubmed",
                    "id": ",".join(ids),
                    "retmode": "xml",
                    "rettype": "abstract",
                    **common,
                },
            )
            records = self._parse(xml_text)
        return SourceSearchResult(
            source=self.name,
            records=records[:limit],
            query=bounded,
            elapsed_seconds=time.monotonic() - started,
        )

    def _parse(self, xml_text: str) -> list[PaperRecord]:
        root = ET.fromstring(xml_text)
        records: list[PaperRecord] = []
        for node in root.findall(".//PubmedArticle"):
            citation = node.find("MedlineCitation")
            article = citation.find("Article") if citation is not None else None
            if citation is None or article is None:
                continue
            pmid = _text(citation.find("PMID"))
            title = _text(article.find("ArticleTitle"))
            authors = []
            for author in article.findall("./AuthorList/Author"):
                collective = _text(author.find("CollectiveName"))
                personal = " ".join(
                    part for part in (
                        _text(author.find("ForeName")),
                        _text(author.find("LastName")),
                    ) if part
                )
                if collective or personal:
                    authors.append(collective or personal)
            abstract_parts = []
            for item in article.findall("./Abstract/AbstractText"):
                content = _text(item)
                label = item.attrib.get("Label")
                abstract_parts.append(f"{label}: {content}" if label and content else content)
            journal = _text(article.find("./Journal/Title"))
            volume = _text(article.find("./Journal/JournalIssue/Volume"))
            issue = _text(article.find("./Journal/JournalIssue/Issue"))
            pages = _text(article.find("Pagination/MedlinePgn"))
            pub_date = article.find("./Journal/JournalIssue/PubDate")
            pub_year = _year(_text(pub_date.find("Year")) if pub_date is not None else "")
            if pub_year is None and pub_date is not None:
                pub_year = _year(_text(pub_date.find("MedlineDate")))
            ids = {
                item.attrib.get("IdType", "").lower(): _text(item)
                for item in node.findall("./PubmedData/ArticleIdList/ArticleId")
            }
            mesh_terms = []
            for heading in citation.findall("./MeshHeadingList/MeshHeading"):
                descriptor = _text(heading.find("DescriptorName"))
                qualifiers = [_text(item) for item in heading.findall("QualifierName")]
                if descriptor:
                    mesh_terms.append(
                        f"{descriptor} / {', '.join(item for item in qualifiers if item)}"
                        if any(qualifiers) else descriptor
                    )
            keywords = [
                _text(item) for item in citation.findall("./KeywordList/Keyword") if _text(item)
            ]
            publication_types = [
                _text(item) for item in article.findall("./PublicationTypeList/PublicationType")
                if _text(item)
            ]
            pmcid = ids.get("pmc", "").upper()
            is_retracted = any(
                "retract" in item.casefold() for item in publication_types
            )
            records.append(PaperRecord(
                title=title,
                source=self.name,
                sources=[self.name],
                authors=authors,
                year=pub_year,
                publication=journal,
                volume=volume,
                issue=issue,
                pages=pages,
                abstract=" ".join(part for part in abstract_parts if part),
                doi=_clean_doi(ids.get("doi")),
                pmid=ids.get("pubmed", pmid),
                pmcid=pmcid,
                url=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else "",
                is_open_access=True if pmcid else None,
                oa_url=f"https://europepmc.org/articles/{pmcid}" if pmcid else "",
                mesh_terms=mesh_terms,
                keywords=keywords,
                publication_types=publication_types,
                is_retracted=is_retracted,
            ))
        return records


class EuropePMCAdapter(SourceAdapter):
    name = "europepmc"
    _url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"

    def __init__(self, http: HTTPClient | None = None) -> None:
        self.http = http or HTTPClient()

    def search(
        self,
        query: str,
        start_year: int | None,
        end_year: int | None,
        limit: int,
        credentials: Mapping[str, str] | None = None,
    ) -> SourceSearchResult:
        started = time.monotonic()
        limit = _bounded_limit(limit)
        bounded = _bounded_query(query, start_year, end_year, self.name)
        payload = self.http.get_json(self._url, {
            "query": bounded,
            "format": "json",
            "resultType": "core",
            "pageSize": limit,
        })
        records = []
        for item in payload.get("resultList", {}).get("result", []):
            authors = [
                author.get("fullName") or " ".join(
                    part for part in (author.get("firstName"), author.get("lastName")) if part
                )
                for author in item.get("authorList", {}).get("author", [])
            ]
            pmcid = str(item.get("pmcid") or "").upper()
            pmid = str(item.get("pmid") or "")
            is_open_access = _bool_or_none(item.get("isOpenAccess"))
            full_text_urls = item.get("fullTextUrlList", {}).get("fullTextUrl", []) or []
            oa_url = ""
            oa_license = ""
            if is_open_access:
                open_urls = [
                    full_text for full_text in full_text_urls
                    if isinstance(full_text, Mapping)
                    and str(full_text.get("availabilityCode") or "").upper() in {"F", "OA"}
                ]
                for full_text in sorted(
                    open_urls,
                    key=lambda value: str(value.get("documentStyle") or "").lower() != "pdf",
                ):
                    candidate = str(full_text.get("url") or "")
                    if candidate:
                        oa_url = candidate
                        oa_license = str(item.get("license") or "")
                        break
                if not oa_url and pmcid:
                    oa_url = f"https://europepmc.org/articles/{pmcid}"
            mesh_terms = []
            for heading in item.get("meshHeadingList", {}).get("meshHeading", []) or []:
                if isinstance(heading, Mapping):
                    term = heading.get("descriptorName") or heading.get("majorTopic_YN")
                    if isinstance(term, Mapping):
                        term = term.get("value")
                    if term:
                        mesh_terms.append(str(term))
            keywords = item.get("keywordList", {}).get("keyword", []) or []
            publication_types = item.get("pubTypeList", {}).get("pubType", []) or []
            records.append(PaperRecord(
                title=_strip_markup(item.get("title")),
                source=self.name,
                sources=[self.name],
                authors=[author for author in authors if author],
                year=_year(item.get("pubYear") or item.get("firstPublicationDate")),
                publication=str(item.get("journalTitle") or ""),
                volume=str(item.get("journalVolume") or ""),
                issue=str(item.get("issue") or ""),
                pages=str(item.get("pageInfo") or ""),
                abstract=_strip_markup(item.get("abstractText")),
                doi=_clean_doi(item.get("doi")),
                pmid=pmid,
                pmcid=pmcid,
                url=(f"https://europepmc.org/article/MED/{pmid}" if pmid else
                     f"https://europepmc.org/article/PMC/{pmcid}" if pmcid else ""),
                citations=_int_or_none(item.get("citedByCount")),
                is_open_access=is_open_access,
                oa_url=oa_url,
                oa_license=oa_license,
                mesh_terms=[str(term) for term in mesh_terms if term],
                keywords=[str(term) for term in keywords if term],
                publication_types=[str(term) for term in publication_types if term],
                is_retracted=_bool_or_none(item.get("isRetracted")) is True,
            ))
        return SourceSearchResult(
            source=self.name,
            records=records[:limit],
            query=bounded,
            elapsed_seconds=time.monotonic() - started,
        )


class OpenAlexAdapter(SourceAdapter):
    name = "openalex"
    _url = "https://api.openalex.org/works"

    def __init__(self, http: HTTPClient | None = None) -> None:
        self.http = http or HTTPClient()

    def search(
        self,
        query: str,
        start_year: int | None,
        end_year: int | None,
        limit: int,
        credentials: Mapping[str, str] | None = None,
    ) -> SourceSearchResult:
        started = time.monotonic()
        credentials = credentials or {}
        limit = _bounded_limit(limit)
        api_key = credentials.get("openalex_api_key")
        if not api_key:
            raise ValueError("OpenAlex requires credentials['openalex_api_key']")
        filters = []
        if start_year is not None:
            filters.append(f"from_publication_date:{start_year}-01-01")
        if end_year is not None:
            filters.append(f"to_publication_date:{end_year}-12-31")
        params: dict[str, Any] = {
            "search": query,
            "per-page": limit,
            "api_key": api_key,
            "select": (
                "id,doi,title,display_name,authorships,publication_year,publication_date,"
                "ids,primary_location,best_oa_location,open_access,abstract_inverted_index,"
                "cited_by_count,is_retracted,biblio,type"
            ),
        }
        if filters:
            params["filter"] = ",".join(filters)
        if credentials.get("email"):
            params["mailto"] = credentials["email"]
        payload = self.http.get_json(self._url, params)
        records = []
        for item in payload.get("results", []):
            primary = item.get("primary_location") or {}
            source = primary.get("source") or {}
            ids = item.get("ids") or {}
            oa = item.get("open_access") or {}
            best_oa = item.get("best_oa_location") or {}
            authors = []
            for authorship in item.get("authorships") or []:
                name = (authorship.get("author") or {}).get("display_name")
                if name:
                    authors.append(name)
            doi = _clean_doi(ids.get("doi") or item.get("doi"))
            openalex_id = _clean_openalex_id(ids.get("openalex") or item.get("id"))
            records.append(PaperRecord(
                title=_strip_markup(item.get("display_name") or item.get("title")),
                source=self.name,
                sources=[self.name],
                authors=authors,
                year=_year(item.get("publication_year") or item.get("publication_date")),
                publication=str(source.get("display_name") or ""),
                volume=str(item.get("biblio", {}).get("volume") or ""),
                issue=str(item.get("biblio", {}).get("issue") or ""),
                pages="-".join(
                    str(value) for value in (
                        item.get("biblio", {}).get("first_page"),
                        item.get("biblio", {}).get("last_page"),
                    ) if value
                ),
                abstract=_openalex_abstract(item.get("abstract_inverted_index")),
                doi=doi,
                openalex_id=openalex_id,
                url=str(primary.get("landing_page_url") or item.get("id") or ""),
                citations=_int_or_none(item.get("cited_by_count")),
                is_open_access=_bool_or_none(oa.get("is_oa")),
                oa_url=str(
                    best_oa.get("pdf_url") or best_oa.get("landing_page_url") or ""
                ) if oa.get("is_oa") else "",
                oa_license=str(best_oa.get("license") or "") if oa.get("is_oa") else "",
                publication_types=[str(item.get("type") or "")] if item.get("type") else [],
                is_retracted=bool(item.get("is_retracted")),
            ))
        return SourceSearchResult(
            source=self.name,
            records=records[:limit],
            query=query,
            elapsed_seconds=time.monotonic() - started,
        )


class CrossrefAdapter(SourceAdapter):
    name = "crossref"
    _url = "https://api.crossref.org/works"

    def __init__(self, http: HTTPClient | None = None) -> None:
        self.http = http or HTTPClient()

    def search(
        self,
        query: str,
        start_year: int | None,
        end_year: int | None,
        limit: int,
        credentials: Mapping[str, str] | None = None,
    ) -> SourceSearchResult:
        started = time.monotonic()
        credentials = credentials or {}
        limit = _bounded_limit(limit)
        filters = []
        if start_year is not None:
            filters.append(f"from-pub-date:{start_year}-01-01")
        if end_year is not None:
            filters.append(f"until-pub-date:{end_year}-12-31")
        params: dict[str, Any] = {
            "query.bibliographic": query,
            "rows": limit,
            "mailto": credentials.get("email"),
        }
        if filters:
            params["filter"] = ",".join(filters)
        payload = self.http.get_json(self._url, params)
        records = []
        for item in payload.get("message", {}).get("items", []):
            title = next(iter(item.get("title") or []), "")
            container = next(iter(item.get("container-title") or []), "")
            authors = []
            for author in item.get("author") or []:
                name = " ".join(
                    part for part in (author.get("given"), author.get("family")) if part
                )
                if name:
                    authors.append(name)
            published = item.get("published-print") or item.get("published-online") or item.get("issued") or {}
            date_parts = published.get("date-parts") or []
            published_year = _year(date_parts[0][0]) if date_parts and date_parts[0] else None
            doi = _clean_doi(item.get("DOI"))
            records.append(PaperRecord(
                title=_strip_markup(title),
                source=self.name,
                sources=[self.name],
                authors=authors,
                year=published_year,
                publication=str(container),
                volume=str(item.get("volume") or ""),
                issue=str(item.get("issue") or ""),
                pages=str(item.get("page") or item.get("article-number") or ""),
                abstract=_strip_markup(item.get("abstract")),
                doi=doi,
                url=str(item.get("URL") or (f"https://doi.org/{doi}" if doi else "")),
                citations=_int_or_none(item.get("is-referenced-by-count")),
                publication_types=[str(item.get("type") or "")] if item.get("type") else [],
                is_retracted=str(item.get("subtype") or "").casefold() == "retracted",
            ))
        return SourceSearchResult(
            source=self.name,
            records=records[:limit],
            query=query,
            elapsed_seconds=time.monotonic() - started,
        )


def _bounded_query(
    query: str,
    start_year: int | None,
    end_year: int | None,
    source: str,
) -> str:
    if start_year is None and end_year is None:
        return query
    lower = start_year if start_year is not None else 1800
    upper = end_year if end_year is not None else 2100
    if source == "pubmed":
        return f"({query}) AND ({lower}:{upper}[pdat])"
    if source == "europepmc":
        return f"({query}) AND FIRST_PDATE:[{lower}-01-01 TO {upper}-12-31]"
    return query


def _openalex_abstract(index: Any) -> str:
    if not isinstance(index, Mapping):
        return ""
    words = []
    for word, positions in index.items():
        if isinstance(positions, list):
            words.extend((position, str(word)) for position in positions if isinstance(position, int))
    return " ".join(word for _, word in sorted(words))


def _int_or_none(value: Any) -> int | None:
    try:
        return int(value) if value is not None and value != "" else None
    except (TypeError, ValueError):
        return None


def _bool_or_none(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if str(value).lower() in {"true", "y", "yes", "1"}:
        return True
    if str(value).lower() in {"false", "n", "no", "0"}:
        return False
    return None


SOURCE_REGISTRY: dict[str, type[SourceAdapter]] = {
    "pubmed": PubMedAdapter,
    "europepmc": EuropePMCAdapter,
    "openalex": OpenAlexAdapter,
    "crossref": CrossrefAdapter,
}


_REWRITE_SOURCE_KEYS = ("pubmed", "europe_pmc", "openalex", "crossref")
_SOURCE_TO_REWRITE_KEY = {
    "pubmed": "pubmed",
    "europepmc": "europe_pmc",
    "europe_pmc": "europe_pmc",
    "openalex": "openalex",
    "crossref": "crossref",
}


def rewrite_search_queries(
    query: str,
    llm_callback: Callable[[str], Any] | None = None,
    enabled: bool = True,
) -> dict[str, Any]:
    """Return one deterministic query per source, falling back atomically."""
    original = query.strip()
    if not original:
        return {
            **{key: "" for key in _REWRITE_SOURCE_KEYS},
            "terms_en": [],
            "terms_zh": [],
        }
    fallback = {
        **{key: original for key in _REWRITE_SOURCE_KEYS},
        "terms_en": [],
        "terms_zh": [],
    }
    if not enabled or llm_callback is None:
        return fallback
    prompt = (
        "Rewrite the scholarly literature search for four databases. Preserve "
        "technical terms, expand useful synonyms, and do not invent dates, IDs, "
        "papers, or other constraints. Return JSON only with exactly these keys: "
        '{"pubmed":"...","europe_pmc":"...","openalex":"...",'
        '"crossref":"...","terms_en":["..."],"terms_zh":["..."]}.\n\n'
        f"Search: {original}"
    )
    try:
        response = llm_callback(prompt)
        parsed = _rewrite_from_json(response)
    except Exception:
        return fallback
    if parsed is None:
        return fallback
    rewritten: dict[str, Any] = {}
    for key in _REWRITE_SOURCE_KEYS:
        value = re.sub(r"\s+", " ", str(parsed.get(key) or "")).strip()
        rewritten[key] = value or original
    for key in ("terms_en", "terms_zh"):
        values = parsed.get(key, [])
        if isinstance(values, str):
            values = [values]
        rewritten[key] = list(dict.fromkeys(
            re.sub(r"\s+", " ", str(value)).strip()
            for value in values if str(value).strip()
        )) if isinstance(values, list) else []
    return rewritten


def _rewrite_from_json(response: Any) -> dict[str, Any] | None:
    parsed = _json_response(response)
    if not isinstance(parsed, Mapping):
        return None
    if not all(key in parsed for key in _REWRITE_SOURCE_KEYS):
        return None
    return dict(parsed)


def _json_response(response: Any) -> Any:
    if isinstance(response, Mapping):
        if "choices" in response:
            choices = response.get("choices") or []
            if choices:
                message = choices[0].get("message", {}) if isinstance(choices[0], Mapping) else {}
                response = message.get("content", "")
        elif "content" in response and not all(key in response for key in _REWRITE_SOURCE_KEYS):
            response = response.get("content")
        else:
            return response
    if not isinstance(response, str):
        return response
    text = response.strip()
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.I)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"(?:\{.*\}|\[.*\])", text, flags=re.S)
        if not match:
            return None
        try:
            return json.loads(match.group())
        except json.JSONDecodeError:
            return None


def _queries_from_json(response: Any) -> list[str]:
    """Parse the legacy generic-query response shape for external callers."""
    if isinstance(response, Mapping):
        if "choices" in response:
            choices = response.get("choices") or []
            if choices:
                message = choices[0].get("message", {}) if isinstance(choices[0], Mapping) else {}
                response = message.get("content", "")
        elif "content" in response and not any(
            key in response for key in ("queries", "search_queries", "rewritten_queries")
        ):
            response = response.get("content")
    parsed: Any = response
    if isinstance(response, str):
        text = response.strip()
        text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.I)
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            match = re.search(r"(?:\{.*\}|\[.*\])", text, flags=re.S)
            if not match:
                return []
            try:
                parsed = json.loads(match.group())
            except json.JSONDecodeError:
                return []
    if isinstance(parsed, Mapping):
        parsed = next(
            (parsed.get(key) for key in ("queries", "search_queries", "rewritten_queries") if key in parsed),
            [],
        )
    if isinstance(parsed, str):
        parsed = [parsed]
    return [str(item) for item in parsed if isinstance(item, str)] if isinstance(parsed, list) else []


def search_literature(
    query: str,
    selected_sources: Sequence[str] | None = None,
    start_year: int | None = None,
    end_year: int | None = None,
    limit: int = 20,
    credentials: Mapping[str, Any] | None = None,
    smart_rewrite: bool = False,
    llm_callback: Callable[[str], Any] | None = None,
    adapters: Mapping[str, SourceAdapter | type[SourceAdapter]] | Sequence[SourceAdapter] | None = None,
) -> FederatedSearchResult:
    """Search providers concurrently, isolate failures, merge, and rank results."""
    search_started = time.monotonic()
    if not query or not query.strip():
        raise ValueError("query must not be empty")
    if start_year is not None and end_year is not None and start_year > end_year:
        raise ValueError("start_year must be less than or equal to end_year")
    limit = _bounded_limit(limit)
    resolved = _resolve_adapters(selected_sources, adapters)
    rewrite = rewrite_search_queries(query, llm_callback, smart_rewrite)
    source_queries = {
        name: str(rewrite[_SOURCE_TO_REWRITE_KEY.get(name, name)] or query.strip())
        for name in resolved
    }
    queries = list(dict.fromkeys(source_queries.values()))
    source_results: dict[str, SourceSearchResult] = {}
    ranked_lists: list[list[PaperRecord]] = []
    errors: dict[str, str] = {}

    def run_source(name: str, adapter: SourceAdapter) -> tuple[SourceSearchResult, list[list[PaperRecord]]]:
        started = time.monotonic()
        records: list[PaperRecord] = []
        lists: list[list[PaperRecord]] = []
        failures = []
        source_credentials = _source_credentials(credentials, name)
        rewritten = source_queries[name]
        try:
            result = adapter.search(rewritten, start_year, end_year, limit, source_credentials)
            if result.error:
                failures.append(f"{rewritten}: {result.error}")
            filtered = [
                record for record in result.records
                if _within_year(record.year, start_year, end_year)
            ][:limit]
            fetched_at = datetime.now(timezone.utc).isoformat()
            for rank, record in enumerate(filtered, 1):
                record.rank = rank
                record.score = 1.0 / (60 + rank)
                if not record.source:
                    record.source = name
                if not record.sources:
                    record.sources = [record.source]
                record.source_ranks[name] = rank
                record.source_ids.update(_record_source_ids(record, name))
                if not record.fetched_at:
                    record.fetched_at = fetched_at
                if not record.field_sources:
                    record.field_sources = _initial_field_sources(record)
            lists.append(filtered)
            records.extend(filtered)
        except Exception as exc:
            failures.append(f"{rewritten}: {type(exc).__name__}: {exc}")
        error = "; ".join(failures) or None
        return SourceSearchResult(
            source=name,
            records=records,
            query=rewritten,
            error=error,
            elapsed_seconds=time.monotonic() - started,
        ), lists

    if resolved:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(resolved), 8)) as executor:
            futures = {
                executor.submit(run_source, name, adapter): name
                for name, adapter in resolved.items()
            }
            for future in concurrent.futures.as_completed(futures):
                name = futures[future]
                try:
                    result, lists = future.result()
                except Exception as exc:
                    result = SourceSearchResult(
                        source=name,
                        query=source_queries[name],
                        error=f"{type(exc).__name__}: {exc}",
                    )
                    lists = []
                source_results[name] = result
                ranked_lists.extend(lists)
                if result.error:
                    errors[name] = result.error

    raw_count = sum(len(ranked) for ranked in ranked_lists)
    merged = _deduplicate(record for ranked in ranked_lists for record in ranked)
    merged.sort(
        key=lambda item: (
            item.score,
            _completeness_score(item),
            item.citations or -1,
            item.year or 0,
        ),
        reverse=True,
    )
    final_records = merged[:200]
    for rank, record in enumerate(final_records, 1):
        record.rank = rank
        _apply_quality(record)
    return FederatedSearchResult(
        records=final_records,
        source_results={name: source_results[name] for name in resolved if name in source_results},
        queries=queries,
        original_query=query.strip(),
        source_queries=source_queries,
        terms_en=list(rewrite.get("terms_en", [])),
        terms_zh=list(rewrite.get("terms_zh", [])),
        raw_count=raw_count,
        unique_count=len(merged),
        duplicate_count=max(0, raw_count - len(merged)),
        source_counts={name: len(result.records) for name, result in source_results.items()},
        elapsed_seconds=time.monotonic() - search_started,
        errors=errors,
    )


def _resolve_adapters(
    selected_sources: Sequence[str] | None,
    adapters: Mapping[str, SourceAdapter | type[SourceAdapter]] | Sequence[SourceAdapter] | None,
) -> dict[str, SourceAdapter]:
    supplied: dict[str, SourceAdapter | type[SourceAdapter]] = {}
    if isinstance(adapters, Mapping):
        supplied = {str(name).lower(): adapter for name, adapter in adapters.items()}
    elif adapters is not None:
        supplied = {adapter.name.lower(): adapter for adapter in adapters}
    names = [name.lower() for name in (selected_sources or supplied.keys() or SOURCE_REGISTRY.keys())]
    resolved: dict[str, SourceAdapter] = {}
    for name in dict.fromkeys(names):
        candidate = supplied.get(name, SOURCE_REGISTRY.get(name))
        if candidate is None:
            raise ValueError(f"Unknown literature source: {name}")
        resolved[name] = candidate() if isinstance(candidate, type) else candidate
    return resolved


def _source_credentials(credentials: Mapping[str, Any] | None, source: str) -> dict[str, str]:
    if not credentials:
        return {}
    common = {key: str(value) for key, value in credentials.items() if not isinstance(value, Mapping)}
    nested = credentials.get(source)
    if isinstance(nested, Mapping):
        common.update({key: str(value) for key, value in nested.items() if value is not None})
    return common


def _within_year(year: int | None, start_year: int | None, end_year: int | None) -> bool:
    if year is None:
        return True
    return (start_year is None or year >= start_year) and (end_year is None or year <= end_year)


def _record_source_ids(record: PaperRecord, source: str) -> dict[str, str]:
    identifiers = {
        "doi": record.doi,
        "pmid": record.pmid,
        "pmcid": record.pmcid,
        "openalex_id": record.openalex_id,
    }
    primary = next((value for value in identifiers.values() if value), "")
    return {source: primary} if primary else {}


def _initial_field_sources(record: PaperRecord) -> dict[str, str]:
    return {
        field_name: record.source
        for field_name in (
            "title", "authors", "year", "publication", "volume", "issue", "pages",
            "abstract", "doi", "pmid", "pmcid", "openalex_id", "url", "citations",
            "is_open_access", "oa_url", "oa_license", "mesh_terms", "keywords",
            "publication_types", "is_retracted",
        )
        if getattr(record, field_name) not in (None, "", [], False)
    }


def _deduplicate(records: Iterable[PaperRecord]) -> list[PaperRecord]:
    groups: list[list[PaperRecord]] = []
    identifier_indexes: dict[str, dict[str, int]] = {
        name: {} for name in ("doi", "pmid", "pmcid", "openalex_id")
    }
    exact_index: dict[tuple[str, int | None], list[int]] = {}
    for original in records:
        record = copy.deepcopy(original)
        record.doi = _clean_doi(record.doi)
        record.pmid = str(record.pmid or "").strip()
        record.pmcid = str(record.pmcid or "").strip().upper()
        record.openalex_id = _clean_openalex_id(record.openalex_id)
        title_key = _normalize_title(record.title)
        group_index: int | None = None
        for field_name in ("doi", "pmid", "pmcid", "openalex_id"):
            value = getattr(record, field_name)
            if value and value in identifier_indexes[field_name]:
                group_index = identifier_indexes[field_name][value]
                break
        if group_index is None and title_key:
            for candidate in exact_index.get((title_key, record.year), []):
                if _doi_compatible(record, groups[candidate]):
                    group_index = candidate
                    break
        if group_index is None:
            group_index = _fuzzy_group(record, title_key, groups)
        if group_index is None:
            group_index = len(groups)
            groups.append([])
        groups[group_index].append(record)
        for field_name, index in identifier_indexes.items():
            value = getattr(record, field_name)
            if value:
                index.setdefault(value, group_index)
        if title_key:
            exact_index.setdefault((title_key, record.year), []).append(group_index)
    return [_coalesce(group) for group in groups]


def _fuzzy_group(
    record: PaperRecord,
    title_key: str,
    groups: Sequence[Sequence[PaperRecord]],
) -> int | None:
    if not title_key or record.year is None or not record.authors:
        return None
    author = _normalize_author(record.authors[0])
    if not author:
        return None
    best: tuple[float, int] | None = None
    for index, group in enumerate(groups):
        if not _doi_compatible(record, group):
            continue
        for candidate in group:
            if candidate.year is None or abs(record.year - candidate.year) > 1:
                continue
            if not candidate.authors or _normalize_author(candidate.authors[0]) != author:
                continue
            candidate_title = _normalize_title(candidate.title)
            ratio = difflib.SequenceMatcher(None, title_key, candidate_title).ratio()
            if ratio >= 0.96 and (best is None or ratio > best[0]):
                best = (ratio, index)
    return best[1] if best else None


def _doi_compatible(record: PaperRecord, group: Sequence[PaperRecord]) -> bool:
    dois = {item.doi for item in group if item.doi}
    return not record.doi or not dois or record.doi in dois


def _normalize_title(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "").casefold()
    value = "".join(character for character in value if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def _normalize_author(value: str) -> str:
    value = unicodedata.normalize("NFKD", value or "").casefold()
    return re.sub(r"[^a-z0-9]+", "", value)


def _coalesce(group: Sequence[PaperRecord]) -> PaperRecord:
    if not group:
        raise ValueError("cannot coalesce an empty record group")
    medical = [item for item in group if item.source in {"pubmed", "europepmc"}]
    crossref = [item for item in group if item.source == "crossref"]
    openalex = [item for item in group if item.source == "openalex"]
    result = copy.deepcopy(max(group, key=lambda item: (item.score, len(item.title))))
    result.title = _best_text(group, "title")
    result.authors = list(max(
        medical or group,
        key=lambda item: (len(item.authors), sum(len(author) for author in item.authors)),
    ).authors)
    result.year = _priority_value("year", crossref, group)
    result.publication = _best_text(crossref, "publication") or _best_text(group, "publication")
    result.volume = _priority_value("volume", crossref, group) or ""
    result.issue = _priority_value("issue", crossref, group) or ""
    result.pages = _priority_value("pages", crossref, group) or ""
    result.abstract = _best_text(medical, "abstract") or _best_text(group, "abstract")
    result.doi = _priority_value("doi", crossref, group) or ""
    result.pmid = _priority_value("pmid", medical, group) or ""
    result.pmcid = _priority_value("pmcid", medical, group) or ""
    result.openalex_id = _priority_value("openalex_id", openalex, group) or ""
    result.url = _priority_value("url", crossref, medical, openalex, group) or ""
    citation_values = [item.citations for item in openalex if item.citations is not None]
    if not citation_values:
        citation_values = [item.citations for item in group if item.citations is not None]
    result.citations = max(citation_values) if citation_values else None
    confirmed_oa = [
        item for item in (*openalex, *medical)
        if item.is_open_access is True and item.oa_url
    ]
    if confirmed_oa:
        result.is_open_access = True
    else:
        result.is_open_access = _priority_value("is_open_access", openalex, medical, group)
    result.oa_url = _priority_value("oa_url", confirmed_oa, openalex, medical, group) or ""
    result.oa_license = _priority_value(
        "oa_license", confirmed_oa, openalex, medical, group
    ) or ""
    if result.oa_url:
        result.is_open_access = True
    result.mesh_terms = list(dict.fromkeys(
        term for item in (medical or group) for term in item.mesh_terms if term
    ))
    result.keywords = list(dict.fromkeys(
        term for item in (medical or group) for term in item.keywords if term
    ))
    result.publication_types = list(dict.fromkeys(
        term for item in group for term in item.publication_types if term
    ))
    result.is_retracted = any(item.is_retracted for item in group)
    result.score = sum(item.score for item in group)
    result.sources = list(dict.fromkeys(
        source for item in group for source in (item.sources or [item.source]) if source
    ))
    result.source = result.sources[0] if result.sources else result.source
    result.source_ids = {
        key: value for item in group for key, value in item.source_ids.items() if value
    }
    result.source_ranks = {
        key: min(
            value,
            *(other.source_ranks.get(key, value) for other in group),
        )
        for item in group for key, value in item.source_ranks.items()
    }
    result.fetched_at = max((item.fetched_at for item in group if item.fetched_at), default="")
    field_sources: dict[str, str] = {}
    selected_values = {
        "title": result.title, "authors": result.authors, "year": result.year,
        "publication": result.publication, "volume": result.volume, "issue": result.issue,
        "pages": result.pages, "abstract": result.abstract, "doi": result.doi,
        "pmid": result.pmid, "pmcid": result.pmcid, "openalex_id": result.openalex_id,
        "url": result.url, "citations": result.citations,
        "is_open_access": result.is_open_access, "oa_url": result.oa_url,
        "oa_license": result.oa_license,
    }
    for field_name, selected in selected_values.items():
        for item in group:
            if getattr(item, field_name) == selected and selected not in (None, "", []):
                field_sources[field_name] = item.field_sources.get(field_name, item.source)
                break
    result.field_sources = field_sources
    conflicts = []
    for field_name in ("doi", "pmid", "pmcid", "openalex_id"):
        if len({getattr(item, field_name) for item in group if getattr(item, field_name)}) > 1:
            conflicts.append(f"conflicting_{field_name}")
    result.quality_flags = conflicts
    return result


def _best_text(records: Sequence[PaperRecord], field_name: str) -> str:
    values = [str(getattr(item, field_name) or "").strip() for item in records]
    return max(values, key=len, default="")


def _first_value(records: Sequence[PaperRecord], field_name: str) -> Any:
    return next((getattr(item, field_name) for item in records if getattr(item, field_name) not in (None, "")), None)


def _priority_value(field_name: str, *record_groups: Sequence[PaperRecord]) -> Any:
    for records in record_groups:
        value = _first_value(records, field_name)
        if value not in (None, "", []):
            return value
    return None


def _quality_flags(record: PaperRecord) -> list[str]:
    flags = list(record.quality_flags)
    for field_name in ("title", "authors", "year", "abstract", "doi"):
        if not getattr(record, field_name):
            flags.append(f"missing_{field_name}")
    if len(record.sources) < 2:
        flags.append("single_source")
    if record.is_open_access is False:
        flags.append("not_open_access")
    if record.is_retracted:
        flags.append("retracted")
    return list(dict.fromkeys(flags))


def _completeness_score(record: PaperRecord) -> int:
    values = (
        record.title, record.authors, record.year, record.publication, record.abstract,
        record.doi or record.pmid, record.oa_url, record.citations,
    )
    return sum(value not in (None, "", []) for value in values)


def _apply_quality(record: PaperRecord) -> None:
    required = ("title", "authors", "year", "publication", "abstract", "doi")
    record.missing_fields = [
        field_name for field_name in required if not getattr(record, field_name)
    ]
    record.summary_ready = bool(record.title and len(record.abstract.strip()) >= 100)
    record.reference_ready = bool(
        record.title and record.authors and record.year and record.publication
        and (record.doi or record.pmid)
    )
    record.needs_enrich = not (record.summary_ready and record.reference_ready)
    record.quality_flags = _quality_flags(record)


def export_excel(data: FederatedSearchResult | Sequence[PaperRecord]) -> bytes:
    """Export records to a formatted XLSX workbook."""
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = [_export_row(record) for record in _records(data)]
    frame = pd.DataFrame(rows, columns=[
        "Rank", "Title", "Authors", "Year", "Publication", "Volume", "Issue",
        "Pages", "Abstract", "DOI", "PMID", "PMCID", "OpenAlex ID", "URL",
        "OA URL", "OA License", "Citations", "Open Access", "Score", "Sources",
        "Summary Ready", "Reference Ready", "Needs Enrich", "Missing Fields",
        "Quality Flags",
    ])
    source_rows = []
    query_rows = []
    if isinstance(data, FederatedSearchResult):
        for name, result in data.source_results.items():
            source_rows.append({
                "Source": name,
                "Raw Count": len(result.records),
                "Query": result.query,
                "Error": result.error or "",
                "Elapsed Seconds": result.elapsed_seconds,
            })
        query_rows = [
            {"Source": name, "Query": query}
            for name, query in data.source_queries.items()
        ]
        query_rows.extend((
            {"Source": "terms_en", "Query": "; ".join(data.terms_en)},
            {"Source": "terms_zh", "Query": "; ".join(data.terms_zh)},
        ))
    source_frame = pd.DataFrame(
        source_rows,
        columns=["Source", "Raw Count", "Query", "Error", "Elapsed Seconds"],
    )
    query_frame = pd.DataFrame(query_rows, columns=["Source", "Query"])
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Papers", index=False)
        source_frame.to_excel(writer, sheet_name="Source_Summary", index=False)
        query_frame.to_excel(writer, sheet_name="Search_Queries", index=False)
        sheet = writer.sheets["Papers"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        widths = {"A": 8, "B": 55, "C": 38, "D": 10, "E": 32, "F": 12,
                  "G": 12, "H": 14, "I": 70, "J": 28, "K": 16, "L": 16,
                  "M": 18, "N": 45, "O": 45, "P": 18, "Q": 12, "R": 13,
                  "S": 12, "T": 24, "U": 15, "V": 17, "W": 14, "X": 30,
                  "Y": 32}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    return output.getvalue()


def export_ris(data: FederatedSearchResult | Sequence[PaperRecord]) -> bytes:
    """Export records as UTF-8 RIS bytes."""
    chunks = []
    for record in _records(data):
        lines = ["TY  - JOUR", f"TI  - {_ris_value(record.title)}"]
        lines.extend(f"AU  - {_ris_value(author)}" for author in record.authors)
        if record.year:
            lines.append(f"PY  - {record.year}")
        if record.publication:
            lines.append(f"JO  - {_ris_value(record.publication)}")
        if record.abstract:
            lines.append(f"AB  - {_ris_value(record.abstract)}")
        if record.doi:
            lines.append(f"DO  - {record.doi}")
        if record.url:
            lines.append(f"UR  - {_ris_value(record.url)}")
        if record.pmid:
            lines.append(f"AN  - PMID:{record.pmid}")
        if record.pmcid:
            lines.append(f"N1  - PMCID:{record.pmcid}")
        lines.append("ER  -")
        chunks.append("\r\n".join(lines))
    return ("\r\n\r\n".join(chunks) + ("\r\n" if chunks else "")).encode("utf-8")


def export_bibtex(data: FederatedSearchResult | Sequence[PaperRecord]) -> bytes:
    """Export records as UTF-8 BibTeX bytes."""
    entries = []
    used_keys: set[str] = set()
    for index, record in enumerate(_records(data), 1):
        key = _bib_key(record, index)
        base, suffix = key, 2
        while key in used_keys:
            key = f"{base}{suffix}"
            suffix += 1
        used_keys.add(key)
        fields = [("title", record.title)]
        if record.authors:
            fields.append(("author", " and ".join(record.authors)))
        fields.extend((
            ("year", str(record.year or "")),
            ("journal", record.publication),
            ("abstract", record.abstract),
            ("doi", record.doi),
            ("url", record.url),
            ("pmid", record.pmid),
            ("pmcid", record.pmcid),
        ))
        rendered = [f"@article{{{key},"]
        rendered.extend(
            f"  {name} = {{{_bib_escape(value)}}},"
            for name, value in fields if value
        )
        rendered.append("}")
        entries.append("\n".join(rendered))
    return ("\n\n".join(entries) + ("\n" if entries else "")).encode("utf-8")


def _records(data: FederatedSearchResult | Sequence[PaperRecord]) -> list[PaperRecord]:
    return list(data.records if isinstance(data, FederatedSearchResult) else data)


def _export_row(record: PaperRecord) -> dict[str, Any]:
    return {
        "Rank": record.rank, "Title": record.title, "Authors": "; ".join(record.authors),
        "Year": record.year, "Publication": record.publication, "Volume": record.volume,
        "Issue": record.issue, "Pages": record.pages, "Abstract": record.abstract,
        "DOI": record.doi, "PMID": record.pmid, "PMCID": record.pmcid,
        "OpenAlex ID": record.openalex_id, "URL": record.url,
        "OA URL": record.oa_url, "OA License": record.oa_license,
        "Citations": record.citations, "Open Access": record.is_open_access,
        "Score": record.score, "Sources": "; ".join(record.sources),
        "Summary Ready": record.summary_ready, "Reference Ready": record.reference_ready,
        "Needs Enrich": record.needs_enrich,
        "Missing Fields": "; ".join(record.missing_fields),
        "Quality Flags": "; ".join(record.quality_flags),
    }


def _ris_value(value: Any) -> str:
    return re.sub(r"[\r\n]+", " ", str(value)).strip()


def _bib_escape(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text.replace("\\", r"\textbackslash{}").replace("{", r"\{").replace("}", r"\}")


def _bib_key(record: PaperRecord, index: int) -> str:
    surname = record.authors[0].split()[-1] if record.authors else "paper"
    token = re.sub(r"[^A-Za-z0-9]", "", surname)
    title_word = next(iter(re.findall(r"[A-Za-z0-9]+", record.title)), "record")
    return f"{token or 'paper'}{record.year or 'nd'}{title_word}{index}"
