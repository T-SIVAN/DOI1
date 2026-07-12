import copy
import io
import json
import re
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from literature_search import (
    EuropePMCAdapter,
    FederatedSearchResult,
    OpenAlexAdapter,
    PaperRecord,
    PubMedAdapter,
    SourceAdapter,
    SourceSearchResult,
    export_bibtex,
    export_excel,
    export_ris,
    rewrite_search_queries,
    search_literature,
)


class _StaticAdapter(SourceAdapter):
    def __init__(self, name, records=None, failure=None):
        self.name = name
        self.records = records or []
        self.failure = failure
        self.calls = []

    def search(self, query, start_year, end_year, limit, credentials=None):
        self.calls.append((query, start_year, end_year, limit, dict(credentials or {})))
        if self.failure:
            raise self.failure
        return SourceSearchResult(
            source=self.name,
            records=copy.deepcopy(self.records),
            query=query,
        )


class LiteratureSearchTests(unittest.TestCase):
    def test_rewrite_per_source_json_and_invalid_fallback(self):
        payload = {
            "pubmed": "base editing[Title/Abstract]",
            "europe_pmc": "TITLE_ABS:\"base editing\"",
            "openalex": "base editing CRISPR",
            "crossref": "base editing",
            "terms_en": ["base editing", "CRISPR"],
            "terms_zh": ["base editing zh"],
        }
        rewritten = rewrite_search_queries(
            "original query",
            lambda _prompt: f"```json\n{json.dumps(payload, ensure_ascii=False)}\n```",
            enabled=True,
        )
        self.assertEqual(rewritten, payload)

        fallback = rewrite_search_queries(
            "original query", lambda _prompt: "not valid JSON", enabled=True
        )
        for source in ("pubmed", "europe_pmc", "openalex", "crossref"):
            self.assertEqual(fallback[source], "original query")
        self.assertEqual(fallback["terms_en"], [])
        self.assertEqual(fallback["terms_zh"], [])

    def test_selected_adapters_called_once_and_partial_failure_isolated(self):
        good = _StaticAdapter(
            "pubmed",
            [PaperRecord(title="Recovered paper", source="pubmed", pmid="123")],
        )
        bad = _StaticAdapter("crossref", failure=RuntimeError("temporary outage"))

        result = search_literature(
            "base editing",
            selected_sources=["pubmed", "crossref"],
            limit=7,
            adapters={"pubmed": good, "crossref": bad},
        )

        self.assertEqual(len(good.calls), 1)
        self.assertEqual(len(bad.calls), 1)
        self.assertEqual(result.raw_count, 1)
        self.assertEqual([record.title for record in result.records], ["Recovered paper"])
        self.assertIn("crossref", result.errors)
        self.assertIn("temporary outage", result.errors["crossref"])
        self.assertIsNone(result.source_results["pubmed"].error)

    def test_doi_dedup_uses_crossref_publication_and_pubmed_abstract(self):
        title = "A precise base editing method"
        shared_doi = "10.1000/shared"
        crossref = _StaticAdapter(
            "crossref",
            [PaperRecord(
                title=title,
                source="crossref",
                authors=["A Lee"],
                year=2024,
                publication="Crossref Journal",
                doi=shared_doi,
            ), PaperRecord(
                title="A second matched paper",
                source="crossref",
                authors=["B Chen"],
                year=2023,
                publication="Crossref Fallback Journal",
            )],
        )
        pubmed = _StaticAdapter(
            "pubmed",
            [PaperRecord(
                title=title,
                source="pubmed",
                authors=["A Lee"],
                year=2024,
                publication="PubMed Journal",
                abstract="PubMed abstract " + ("with clinical detail " * 8),
                doi=shared_doi,
                pmid="111",
            ), PaperRecord(
                title="A second matched paper",
                source="pubmed",
                authors=["B Chen"],
                year=2023,
                publication="PubMed Fallback Journal",
                doi="10.1000/fallback",
                pmid="333",
            )],
        )
        distinct = _StaticAdapter(
            "europepmc",
            [PaperRecord(
                title=title,
                source="europepmc",
                authors=["A Lee"],
                year=2024,
                publication="Other Journal",
                abstract="A separate record",
                doi="10.1000/distinct",
                pmid="222",
            )],
        )

        result = search_literature(
            "base editing",
            selected_sources=["crossref", "pubmed", "europepmc"],
            adapters={
                "crossref": crossref,
                "pubmed": pubmed,
                "europepmc": distinct,
            },
        )

        self.assertEqual(result.raw_count, 5)
        self.assertEqual(result.unique_count, 3)
        self.assertEqual(result.duplicate_count, 2)
        self.assertEqual({record.doi for record in result.records}, {
            shared_doi, "10.1000/distinct", "10.1000/fallback"
        })
        merged = next(record for record in result.records if record.doi == shared_doi)
        self.assertEqual(merged.publication, "Crossref Journal")
        self.assertTrue(merged.abstract.startswith("PubMed abstract"))
        self.assertEqual(set(merged.sources), {"crossref", "pubmed"})

    def test_pubmed_search_makes_two_http_calls_and_parses_xml(self):
        xml_fixture = """\
<PubmedArticleSet>
  <PubmedArticle>
    <MedlineCitation>
      <PMID>12345</PMID>
      <Article>
        <ArticleTitle>Base <i>editing</i> study</ArticleTitle>
        <Abstract><AbstractText Label="BACKGROUND">Detailed result.</AbstractText></Abstract>
        <AuthorList><Author><ForeName>Jane</ForeName><LastName>Doe</LastName></Author></AuthorList>
        <Journal>
          <Title>Genome Medicine</Title>
          <JournalIssue>
            <Volume>12</Volume><Issue>3</Issue><PubDate><Year>2024</Year></PubDate>
          </JournalIssue>
        </Journal>
        <Pagination><MedlinePgn>10-20</MedlinePgn></Pagination>
        <PublicationTypeList><PublicationType>Journal Article</PublicationType></PublicationTypeList>
      </Article>
      <MeshHeadingList>
        <MeshHeading><DescriptorName>Gene Editing</DescriptorName></MeshHeading>
      </MeshHeadingList>
      <KeywordList><Keyword>CRISPR</Keyword></KeywordList>
    </MedlineCitation>
    <PubmedData>
      <ArticleIdList>
        <ArticleId IdType="pubmed">12345</ArticleId>
        <ArticleId IdType="doi">10.1000/TEST</ArticleId>
        <ArticleId IdType="pmc">PMC999</ArticleId>
      </ArticleIdList>
    </PubmedData>
  </PubmedArticle>
</PubmedArticleSet>"""

        class FakeHTTP:
            def __init__(self):
                self.calls = []

            def get_json(self, url, params=None, headers=None):
                self.calls.append(("json", url, params))
                return {"esearchresult": {"idlist": ["12345"]}}

            def get_text(self, url, params=None, headers=None):
                self.calls.append(("text", url, params))
                return xml_fixture

        http = FakeHTTP()
        with patch("literature_search._wait_for_ncbi") as wait_for_ncbi:
            result = PubMedAdapter(http=http).search(
                "base editing", 2024, 2025, 5, {"email": "test@example.org"}
            )

        self.assertEqual(len(http.calls), 2)
        self.assertEqual([call[0] for call in http.calls], ["json", "text"])
        self.assertTrue(http.calls[0][1].endswith("/esearch.fcgi"))
        self.assertTrue(http.calls[1][1].endswith("/efetch.fcgi"))
        self.assertEqual(wait_for_ncbi.call_count, 2)
        self.assertEqual(len(result.records), 1)
        record = result.records[0]
        self.assertEqual(record.title, "Base editing study")
        self.assertEqual(record.authors, ["Jane Doe"])
        self.assertEqual(record.year, 2024)
        self.assertEqual(record.abstract, "BACKGROUND: Detailed result.")
        self.assertEqual(record.doi, "10.1000/test")
        self.assertEqual(record.pmid, "12345")
        self.assertEqual(record.pmcid, "PMC999")
        self.assertEqual(record.mesh_terms, ["Gene Editing"])

    def test_openalex_missing_key_raises_before_http(self):
        class NeverHTTP:
            def __init__(self):
                self.calls = 0

            def get_json(self, *args, **kwargs):
                self.calls += 1
                raise AssertionError("HTTP must not be called without an API key")

        http = NeverHTTP()
        with self.assertRaisesRegex(ValueError, "OpenAlex requires"):
            OpenAlexAdapter(http=http).search("base editing", None, None, 20, {})
        self.assertEqual(http.calls, 0)

    def test_europepmc_preserves_relevance_and_uses_confirmed_oa_url(self):
        class FakeHTTP:
            def __init__(self):
                self.params = None

            def get_json(self, url, params=None, headers=None):
                self.params = params
                return {
                    "resultList": {
                        "result": [{
                            "id": "123",
                            "pmid": "123",
                            "title": "Relevant paper",
                            "pubYear": "2024",
                            "isOpenAccess": "Y",
                            "isRetracted": "N",
                            "license": "cc by",
                            "fullTextUrlList": {"fullTextUrl": [
                                {
                                    "availabilityCode": "S",
                                    "documentStyle": "doi",
                                    "url": "https://doi.org/10.1000/test",
                                },
                                {
                                    "availabilityCode": "F",
                                    "documentStyle": "pdf",
                                    "url": "https://europepmc.org/articles/PMC1?pdf=render",
                                },
                            ]},
                        }]
                    }
                }

        http = FakeHTTP()
        record = EuropePMCAdapter(http=http).search("editing", 2022, 2026, 5).records[0]
        self.assertNotIn("sort", http.params)
        self.assertEqual(record.oa_url, "https://europepmc.org/articles/PMC1?pdf=render")
        self.assertEqual(record.oa_license, "cc by")
        self.assertFalse(record.is_retracted)

    def test_excel_export_has_three_expected_sheets(self):
        record = PaperRecord(
            title="Exported paper",
            source="pubmed",
            sources=["pubmed"],
            authors=["Jane Doe"],
            year=2024,
            publication="Genome Medicine",
            doi="10.1000/export",
        )
        data = FederatedSearchResult(
            records=[record],
            source_results={
                "pubmed": SourceSearchResult(
                    source="pubmed", records=[record], query="base editing"
                )
            },
            source_queries={"pubmed": "base editing"},
        )

        workbook = load_workbook(io.BytesIO(export_excel(data)), read_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                ["Papers", "Source_Summary", "Search_Queries"],
            )
        finally:
            workbook.close()

    def test_ris_record_count_and_bibtex_keys_are_unique(self):
        records = [
            PaperRecord(
                title="Shared title",
                source="pubmed",
                authors=["Jane Doe"],
                year=2024,
                publication="Journal One",
                doi="10.1000/one",
            ),
            PaperRecord(
                title="Shared title",
                source="crossref",
                authors=["Jane Doe"],
                year=2024,
                publication="Journal Two",
                doi="10.1000/two",
            ),
        ]

        ris = export_ris(records).decode("utf-8")
        bibtex = export_bibtex(records).decode("utf-8")
        keys = re.findall(r"@article\{([^,]+),", bibtex)

        self.assertEqual(ris.count("TY  - JOUR"), 2)
        self.assertEqual(ris.count("ER  -"), 2)
        self.assertEqual(len(keys), 2)
        self.assertEqual(len(set(keys)), 2)


if __name__ == "__main__":
    unittest.main()
